# -*- coding: utf-8 -*-
"""
Generator contoh file Excel import soal Portal HC KPB.
Format 9 kolom sesuai parser AssessmentAdminController.ImportPackageQuestions:
  Pertanyaan | Opsi A | Opsi B | Opsi C | Opsi D | Jawaban Benar | Elemen Teknis | QuestionType | Rubrik

QuestionType: MultipleChoice (SA, default jika kosong) | MultipleAnswer (MA) | Essay
- SA  : Opsi A-D wajib, Jawaban Benar 1 huruf (A/B/C/D)
- MA  : Opsi A-D wajib, Jawaban Benar huruf dipisah koma (A,C / A,B,D)
- Essay: Opsi A-D + Jawaban Benar dikosongkan, Rubrik WAJIB

Header baris 1 di-skip otomatis oleh importer (parser pakai .Skip(1)).
Skor per soal & MaxCharacters tidak dimasukkan (parser hardcode 10 / 2000).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADERS = ["Pertanyaan", "Opsi A", "Opsi B", "Opsi C", "Opsi D",
           "Jawaban Benar", "Elemen Teknis", "QuestionType", "Rubrik"]

HEADER_FILL = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# ---------------------------------------------------------------------------
# Data soal (tema K3 + operasi kilang / PROTON)
# Tiap baris: (Pertanyaan, A, B, C, D, JawabanBenar, ElemenTeknis, QuestionType, Rubrik)
# ---------------------------------------------------------------------------

SA = [
    ("Alat Pelindung Diri (APD) yang wajib dikenakan saat memasuki area proses kilang adalah?",
     "Helm safety (safety helmet)", "Sandal jepit", "Kaos oblong", "Topi biasa",
     "A", "K3 Dasar", "MultipleChoice", ""),
    ("Fungsi utama Pressure Safety Valve (PSV) pada vessel bertekanan adalah?",
     "Menaikkan tekanan operasi", "Melindungi peralatan dari overpressure", "Mengukur suhu fluida", "Mengatur laju aliran",
     "B", "Peralatan Statis", "MultipleChoice", ""),
    ("Warna pipa untuk fluida pemadam kebakaran (fire water) sesuai standar warna adalah?",
     "Hijau", "Kuning", "Merah", "Biru",
     "C", "K3 Dasar", "MultipleChoice", ""),
    ("Prosedur LOTO (Lock Out Tag Out) bertujuan untuk?",
     "Mempercepat pekerjaan maintenance", "Mengisolasi sumber energi berbahaya saat maintenance", "Menambah kapasitas produksi", "Menghemat konsumsi listrik",
     "B", "K3 Dasar", "MultipleChoice", ""),
    ("Alat ukur yang digunakan untuk memantau tekanan operasi pada sebuah vessel adalah?",
     "Thermometer", "Pressure gauge", "Flow meter", "Level gauge",
     "B", "Instrumentasi", "MultipleChoice", ""),
]

MA = [
    ("Manakah yang termasuk APD wajib saat bekerja di area proses kilang? (pilih semua yang benar)",
     "Helm safety", "Safety shoes", "Sarung tangan (safety gloves)", "Sepatu sandal",
     "A,B,C", "K3 Dasar", "MultipleAnswer", ""),
    ("Manakah yang termasuk jenis alat penukar panas (heat exchanger)?",
     "Shell & tube", "Plate heat exchanger", "Pompa sentrifugal", "Air cooler",
     "A,B,D", "Peralatan Statis", "MultipleAnswer", ""),
    ("Parameter apa saja yang rutin dipantau operator pada pompa sentrifugal?",
     "Tekanan discharge", "Suhu bearing", "Vibrasi", "Warna cat casing",
     "A,B,C", "Rotating Equipment", "MultipleAnswer", ""),
    ("Manakah yang termasuk sumber bahaya (hazard) di area proses kilang?",
     "Gas mudah terbakar", "Tekanan tinggi", "Suhu tinggi", "Suara musik",
     "A,B,C", "K3 Dasar", "MultipleAnswer", ""),
]

ESSAY = [
    ("Jelaskan langkah-langkah prosedur start-up pompa sentrifugal secara aman.",
     "", "", "", "", "", "Rotating Equipment", "Essay",
     "Jawaban harus mencakup: pastikan suction valve terbuka, lakukan priming, cek arah putaran motor, buka discharge valve secara bertahap, dan monitor tekanan serta vibrasi. Skor penuh bila menyebut minimal 4 langkah secara berurutan."),
    ("Uraikan tindakan yang harus dilakukan operator saat terjadi kebocoran gas mudah terbakar di area proses.",
     "", "", "", "", "", "K3 Dasar", "Essay",
     "Jawaban harus mencakup: aktivasi alarm/notifikasi control room, isolasi sumber kebocoran, evakuasi menuju mustering point, hindari sumber api/percikan, dan lapor ke tim emergency. Skor penuh bila menyebut minimal 4 tindakan."),
    ("Jelaskan fungsi dan prinsip kerja kolom distilasi pada proses pengolahan minyak.",
     "", "", "", "", "", "Proses", "Essay",
     "Jawaban harus mencakup: pemisahan fraksi berdasarkan perbedaan titik didih, peran reflux di bagian atas, peran reboiler di bagian bawah, dan gradien suhu top-bottom. Skor penuh bila menyebut minimal 3 konsep."),
]

CAMPURAN = [SA[0], SA[1], MA[0], MA[2], ESSAY[0], ESSAY[1]]


def build(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Question Import"

    # Header
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")

    # Data
    for r, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Lebar kolom
    widths = [55, 24, 24, 24, 24, 14, 20, 16, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"  wrote {path}  ({len(rows)} soal)")


if __name__ == "__main__":
    import os
    here = os.path.dirname(os.path.abspath(__file__))
    print("Generating contoh file import soal...")
    build(os.path.join(here, "Contoh_Soal_SA.xlsx"), SA)
    build(os.path.join(here, "Contoh_Soal_MA.xlsx"), MA)
    build(os.path.join(here, "Contoh_Soal_Essay.xlsx"), ESSAY)
    build(os.path.join(here, "Contoh_Soal_Campuran.xlsx"), CAMPURAN)
    print("Done.")
